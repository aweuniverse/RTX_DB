# -*- coding: utf-8 -*-
"""
STATEMENT OF PURPOSE:
    Run this program will update two SQL tables from the specified SetOfOrders file:
        1), CUST_ORDER table using tabs named "XX-XX ALL" and "OO":
            ~10.40.44 for a specific season: this must be saved in a tab with naming format "XX-XX ALL"
            ~20.52 for all open orders: this must be saved in a tab named "OO"
        2), HFC_ASSOC table using tabs named "XX-XX ASSOC":
            ~10.40.45 for a specific season: this must be saved in a tabl with naming format "XX-XX ASSOC"
    IT'S IMPORTANT that all the tabs in one "SetOfOrder" file are pulled at the same time from PROCOMM
PRE-REQUISITE:
    SQL table STYLE_MASTER and CUSTOMER need to be fully loaded
IMPORTANT:
    sourcefile string needs to be updated to reflect the correct source file location
"""
import os
from openpyxl import load_workbook
import pandas as pd
import numpy as np
from datetime import datetime as dt
import sqlalchemy
import sys
import logging

os.chdir('W:\\Roytex - The Method\\Ping\\ROYTEXDB')
###IMPORTANT: UPDATE BELOW STRING######
###IMPORTANT: UPDATE BELOW STRING######
###IMPORTANT: UPDATE BELOW STRING######
sourcefile = 'DATASOURCE Archive\\SetOfOrders_7.1.2019.xlsx'
###IMPORTANT: UPDATE ABOVE STRING######
###IMPORTANT: UPDATE ABOVE STRING######
###IMPORTANT: UPDATE ABOVE STRING######
engine = sqlalchemy.create_engine("mssql+pyodbc://@sqlDSN")
conn = engine.connect()

wb = load_workbook(sourcefile, read_only=True)
orderTabs = [each for each in wb.sheetnames if 'ALL' in each]
assocTabs = [each for each in wb.sheetnames if 'ASSOC' in each]

def oneSeasonOrder (aFile, aTab):
    pdOrder = pd.read_excel(aFile, sheet_name=aTab, skiprows=1, header=None, usecols=[0, 2, 4, 5, 6, 7, 8, 10, 11, 13, 14, 16, 18, 19, 21, 22], 
                          names=['DIV', 'CUST_NBR', 'CUST_PO', 'GREEN_BAR', 'ORDER_DATE', 'START_SHIP', 'CXL_SHIP', 'SEASON', 'STYLE', 'COLOR', 'SP', 'ORDERED_UNITS', 'SHIPPED_UNITS', 'SHIP_ON_DATE', 'PICK_UNITS', 'UNCONFIRMED'],
                          converters={'DIV':np.int16, 'CUST_NBR': str, 'CUST_PO': str, 'GREEN_BAR': str, 'ORDER_DATE': str, 'START_SHIP': str, 'CXL_SHIP': str, 'SEASON': str, 'STYLE': str, 'COLOR': str, 'SP': np.float64, 'ORDERED_UNITS': np.int64, 'SHIPPED_UNITS': np.int64, 'SHIP_ON_DATE': str, 'PICK_UNITS': np.int64, 'UNCONFIRMED': str})
    
    pdOrder['CUST_NBR'] = pdOrder['CUST_NBR'].apply(lambda x: x.zfill(5))
    pdOrder['CUST_PO'] = pdOrder['CUST_PO'].apply(lambda x: x.zfill(6) if len(x) < 6 else x)
   
#    pdOrder['ORDER_DATE'] = pd.to_datetime(pdOrder['ORDER_DATE'], errors='coerce')
#    pdOrder['START_SHIP'] = pd.to_datetime(pdOrder['START_SHIP'], errors='coerce')
#    pdOrder['CXL_SHIP'] = pd.to_datetime(pdOrder['CXL_SHIP'], errors='coerce')
#    pdOrder['SHIP_ON_DATE'] = pd.to_datetime(pdOrder['SHIP_ON_DATE'], errors='coerce')
    pdOrder['ORDER_DATE'] = pdOrder['ORDER_DATE'].apply(lambda x: dt.strptime(x[:10], '%Y-%m-%d').date())
    pdOrder['START_SHIP'] = pdOrder['START_SHIP'].apply(lambda x: dt.strptime(x[:10], '%Y-%m-%d').date())
    pdOrder['CXL_SHIP'] = pdOrder['CXL_SHIP'].apply(lambda x: dt.strptime(x[:10], '%Y-%m-%d').date())
    pdOrder['SHIP_ON_DATE'] = pdOrder['SHIP_ON_DATE'].apply(lambda x: dt.strptime(x[:10], '%Y-%m-%d').date() if pd.isnull(x) == False else pd.NaT)
    
    pdOrder['SEASON'] = pdOrder['SEASON'].apply(lambda x: x[:2].upper() + '-' + x[-2:])
    pdOrder['UNCONFIRMED'] = pdOrder['UNCONFIRMED'].apply(lambda x: 1 if x == '*' else 0)
    pdOrder['STYLE'] = pdOrder['STYLE'].apply(lambda x: x.zfill(6) if len(x) < 6 else x)
    pdOrder['COLOR'] = pdOrder['COLOR'].apply(lambda x: x.zfill(3))   
    return pdOrder

def multiSeasonOrder ():
#    fileList = [each for each in os.listdir() if each[:17] == 'SOURCE_CUST_ORDER']
#    if len(fileList) >0:
#        for n in range(len(fileList)):
#            if n == 0:
#                multiSeason = oneSeasonOrder(fileList[n])
#            else:
#                addone = oneSeasonOrder(fileList[n])
#                multiSeason = multiSeason.append(addone)
#    wb = load_workbook(sourcefile, read_only=True)
#    orderTabs = [each for each in wb.sheetnames if 'ALL' in each]
    if len(orderTabs) >0:
        for n in range(len(orderTabs)):
            if n == 0:
                multiSeason = oneSeasonOrder(sourcefile, orderTabs[n])
            else:
                addone = oneSeasonOrder(sourcefile, orderTabs[n])
                multiSeason = multiSeason.append(addone)
#        pdOO = pd.read_csv('SOURCE_OO.csv', skiprows=1, header=None, usecols=[4, 10, 12, 16, 17, 19, 20, 21, 22, 23, 24, 25, 26], 
#                   names=['GREEN_BAR', 'STYLE', 'COLOR', 'SHIPPED', 'BAL', 'S1', 'S2', 'S3', 'S4', 'S5', 'S6', 'S7', 'S8'], converters={4: str, 10: str, 12: str})
        try:
            pdOO = pd.read_excel(sourcefile, sheet_name='OO', skiprows=1, header=None, usecols=[4, 10, 12, 16, 17, 19, 20, 21, 22, 23, 24, 25, 26], 
                                 names=['GREEN_BAR', 'STYLE', 'COLOR', 'SHIPPED', 'BAL', 'S1', 'S2', 'S3', 'S4', 'S5', 'S6', 'S7', 'S8'], 
                                 converters={'GREEN_BAR': str, 'STYLE': str, 'COLOR': str})
        except:
            sys.exit('There is no "OO" file to read')
        pdOO['STYLE'] = pdOO['STYLE'].apply(lambda x: x.zfill(6) if len(x) < 6 else x)
        pdOO['COLOR'] = pdOO['COLOR'].apply(lambda x: x.zfill(3))
        newMulti = pd.merge(multiSeason, pdOO[['GREEN_BAR', 'STYLE', 'COLOR', 'BAL']], how='left', on=['GREEN_BAR', 'STYLE', 'COLOR'])
        newMulti['BAL'].fillna(0, inplace=True)
        """
        Orders are categorized below:
            DEACTIVE are orders that have been cancelled but are still showing in the data pulled (*not sure how it's different from orders no longer showing up)
                update all the balances to 0 (size break)
            COMPLETE are orders that we have shipped in full
                update all the balances to 0 (size break)
            PARTIAL are orders that we have shipped but have balances left
                use OO to update the balances but not the orders (size break)
            UNCONFIRMED are orders that are not confirmed (i.e. in HFC)
                use OO to update both balances and orders (size break)
            ACTIVE are customer PO orders that are in the pipeline waiting to ship
                use OO to update both balances and orders (size break)
        """
        newMulti['COMMENT'] = newMulti.apply(lambda row: 'DEACTIVE' if ((row.SHIPPED_UNITS == 0) & (pd.isnull(row.SHIP_ON_DATE) == False) & (row.BAL == 0))
                else 'COMPLETE' if ((row.SHIPPED_UNITS >= row.ORDERED_UNITS) & (pd.isnull(row.SHIP_ON_DATE) == False))
                else 'PARTIAL' if ((row.BAL > 0) & (pd.isnull(row.SHIP_ON_DATE) == False))
                else 'UNCONFIRMED' if row.UNCONFIRMED == 1
                else 'ACTIVE', axis=1)
        
        ###ADD CODE TO IDENTIFY ORDER CATEGORY: UPFRONT, OR OFF-PRICE###
        
        pdOrder = pdOO[(pdOO['SHIPPED'] == 0) & (pdOO['BAL'] > 0)]
        pdOrder = pdOrder[['GREEN_BAR', 'STYLE', 'COLOR', 'S1', 'S2', 'S3', 'S4', 'S5', 'S6', 'S7', 'S8']]
        
        pdPartial = pdOO[(pdOO['SHIPPED'] > 0) & (pdOO['BAL'] > 0)]
        pdPartial = pdPartial[['GREEN_BAR', 'STYLE', 'COLOR', 'S1', 'S2', 'S3', 'S4', 'S5', 'S6', 'S7', 'S8']]

        newMulti.to_sql('#temp_cust_order', con=conn, if_exists='replace', index=False)
        pdOrder.to_sql('#temp_full_order', con=conn, if_exists='replace', index=False)
        pdPartial.to_sql('#temp_partial_order', con=conn, if_exists='replace', index=False)
        trans = conn.begin()
        try:
            conn.execute("""MERGE DBO.CUST_ORDER AS T 
                         USING dbo.#temp_cust_order AS S 
                         ON (T.GREEN_BAR = S.GREEN_BAR and T.STYLE = S.STYLE and T.COLOR = S.COLOR)
                         WHEN MATCHED THEN UPDATE
                         SET T.DIV=S.DIV, T.CUST_NBR=S.CUST_NBR, T.ORDER_DATE=S.ORDER_DATE, T.START_SHIP=S.START_SHIP, T.CXL_SHIP=S.CXL_SHIP, T.CUST_PO=S.CUST_PO, 
                         T.SEASON=S.SEASON, T.SP=S.SP, T.ORDERED_UNITS=S.ORDERED_UNITS, T.SHIPPED_UNITS=S.SHIPPED_UNITS, T.BAL_UNITS=S.BAL, T.PICK_UNITS=S.PICK_UNITS, 
                         T.SHIP_ON_DATE = S.SHIP_ON_DATE, T.UNCONFIRMED = S.UNCONFIRMED, T.COMMENT=S.COMMENT, T.CXL=0 
                         WHEN NOT MATCHED BY TARGET THEN 
                         INSERT (GREEN_BAR, STYLE, COLOR, DIV, CUST_NBR, ORDER_DATE, START_SHIP, CXL_SHIP, CUST_PO, SEASON, SP, ORDERED_UNITS, SHIPPED_UNITS, BAL_UNITS, PICK_UNITS, SHIP_ON_DATE, UNCONFIRMED, COMMENT) VALUES
                         (S.GREEN_BAR, S.STYLE, S.COLOR, S.DIV, S.CUST_NBR, S.ORDER_DATE, S.START_SHIP, S.CXL_SHIP, S.CUST_PO, S.SEASON, S.SP, S.ORDERED_UNITS, S.SHIPPED_UNITS, S.BAL, S.PICK_UNITS, S.SHIP_ON_DATE, S.UNCONFIRMED, S.COMMENT)
                         WHEN NOT MATCHED BY SOURCE AND (T.SEASON IN (select distinct SEASON from #temp_cust_order)) THEN
                         UPDATE SET T.CXL=1, T.BAL_UNITS = 0, T.BAL_S1 = 0, T.BAL_S2 = 0, T.BAL_S3 = 0, T.BAL_S4 = 0, T.BAL_S5 = 0, T.BAL_S6 = 0, T.BAL_S7 = 0, T.BAL_S8 =0;""")
            conn.execute("""UPDATE T
                         SET T.ORDERED_S1 = S.S1, T.ORDERED_S2 = S.S2, T.ORDERED_S3 = S.S3, T.ORDERED_S4 = S.S4, T.ORDERED_S5 = S.S5, T.ORDERED_S6 = S.S6, T.ORDERED_S7 = S.S7, T.ORDERED_S8 = S.S8, 
                         T.BAL_S1 = S.S1, T.BAL_S2 = S.S2, T.BAL_S3 = S.S3, T.BAL_S4 = S.S4, T.BAL_S5 = S.S5, T.BAL_S6 = S.S6, T.BAL_S7 = S.S7, T.BAL_S8 = S.S8
                         FROM DBO.CUST_ORDER AS T INNER JOIN #temp_full_order AS S ON (S.GREEN_BAR = T.GREEN_BAR and S.STYLE = T.STYLE and S.COLOR = T.COLOR);""") 
            conn.execute("""UPDATE T
                         SET T.BAL_S1 = S.S1, T.BAL_S2 = S.S2, T.BAL_S3 = S.S3, T.BAL_S4 = S.S4, T.BAL_S5 = S.S5, T.BAL_S6 = S.S6, T.BAL_S7 = S.S7, T.BAL_S8 = S.S8
                         FROM DBO.CUST_ORDER AS T INNER JOIN #temp_partial_order AS S ON (S.GREEN_BAR = T.GREEN_BAR and S.STYLE = T.STYLE and S.COLOR = T.COLOR);""")
            conn.execute("""UPDATE T
                         SET T.BAL_S1 = 0, T.BAL_S2 = 0, T.BAL_S3 = 0, T.BAL_S4 = 0, T.BAL_S5 = 0, T.BAL_S6 = 0, T.BAL_S7 = 0, T.BAL_S8 = 0
                         FROM DBO.CUST_ORDER AS T WHERE T.COMMENT IN ('COMPLETE', 'DEACTIVE');""")
            trans.commit()
            print ('CUST_ORDER TABLE UPDATE COMPLETED SUCCESSFULLY FROM SOURCE ' + sourcefile.split("\\")[-1])
        except Exception as e:
            logger = logging.Logger('Catch_All')
            logger.error(str(e))
            trans.rollback()
            conn.close()
            engine.dispose()
    else:
        sys.exit('There is no customer order file to read')

def oneSeason (aFile, aTab):
    assoc = pd.read_excel(aFile, sheet_name=aTab, skiprows=1, header=None, names=['GREEN_BAR', 'STYLE', 'COLOR', 'HFC'], usecols=[1, 3, 4, 5], converters={'GREEN_BAR': str, 'STYLE': str, 'COLOR':str, 'HFC':str})
    assoc['STYLE'] = assoc['STYLE'].apply(lambda x: x.zfill(6) if len(x) < 6 else x)
    assoc['COLOR'] = assoc['COLOR'].apply(lambda x: x.zfill(3))
    assoc['HFC'] = assoc['HFC'].apply(lambda x: x.split('-')[0].zfill(6) if pd.isnull(x) == False else x)
    assoc['SEASON'] = aTab[:5].upper()    
    return assoc

def allSeasonAssoc ():
#    fileList = [each for each in os.listdir() if each[:16] == 'SOURCE_HFC_ASSOC']
    if len(assocTabs) >0:
        for n in range(len(assocTabs)):
            if n == 0:
                allAssoc = oneSeason(sourcefile, assocTabs[n])
            else:
                addone = oneSeason(sourcefile, assocTabs[n])
                allAssoc = allAssoc.append(addone)
        allAssoc.to_sql('#temp_hfc_assoc', con=conn, if_exists='replace', index=False)
        trans = conn.begin()
        try:
            conn.execute("""MERGE DBO.HFC_ASSOC AS T 
                         USING dbo.#temp_hfc_assoc AS S 
                         ON (T.GREEN_BAR = S.GREEN_BAR and T.STYLE=S.STYLE and T.COLOR=S.COLOR) 
                         WHEN MATCHED THEN UPDATE
                         SET T.HFC=S.HFC, T.SEASON=S.SEASON, T.CXL=0
                         WHEN NOT MATCHED BY TARGET THEN 
                         INSERT (GREEN_BAR, STYLE, COLOR, HFC, SEASON) VALUES
                         (S.GREEN_BAR, S.STYLE, S.COLOR, S.HFC, S.SEASON)
                         WHEN NOT MATCHED BY SOURCE THEN
                         UPDATE SET T.CXL=1;""")
            trans.commit()         
            print ('HFC_ASSOC TABLE UPDATE COMPLETED SUCCESSFULLY FROM SOURCE ' + sourcefile.split("\\")[-1])
        except Exception as e:
            logger = logging.Logger('Catch_All')
            logger.error(str(e))
            trans.rollback()
            conn.close()
            engine.dispose()
    else:
        sys.exit('There is no association file to read')

multiSeasonOrder()
allSeasonAssoc()

conn.close()
engine.dispose()

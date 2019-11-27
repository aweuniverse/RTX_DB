# -*- coding: utf-8 -*-
"""
STATEMENT OF PURPOSE:
    Run this program will update 3 SQL tables from the specified SetOfOrders file
    The 3 tables can also be updated separately if needed (run each function separately)
        1), CUST_ORDER table using tabs named "XX-XX ALL" and "OO":
            ~10.40.44 for a specific season: this must be saved in a tab with naming format "XX-XX ALL"
            ~20.52 for all open orders: this must be saved in a tab named "OO"
        2), HFC_ASSOC table using tabs named "XX-XX ASSOC":
            ~10.40.45 for a specific season: this must be saved in a tabl with naming format "XX-XX ASSOC"
        3), INV table using tab named "INV"
    IT'S IMPORTANT that all the tabs in one "SetOfOrder" file are pulled at the same time from PROCOMM
PRE-REQUISITE:
    SQL table STYLE_MASTER and CUSTOMER need to be fully loaded
IMPORTANT:
    sourcefile string needs to be updated to reflect the correct source file location
"""
import os
from openpyxl import load_workbook
import pandas as pd
import numpy as np
from datetime import datetime as dt
import sqlalchemy
import sys
import logging

os.chdir('W:\\Roytex - The Method\\Ping\\ROYTEXDB')

sourcefile = 'DATASOURCE Archive\\SetOfOrders_11.22.2019.xlsx'  ###IMPORTANT: UPDATE THIS FILE LOCATION STRING######

engine = sqlalchemy.create_engine("mssql+pyodbc://@sqlDSN")
conn = engine.connect()

wb = load_workbook(sourcefile, read_only=True)
orderTabs = [each for each in wb.sheetnames if 'ALL' in each]
assocTabs = [each for each in wb.sheetnames if 'ASSOC' in each]
invTab = 'INV'

def oneSeasonOrder (aFile, aTab):
    pdOrder = pd.read_excel(aFile, aTab, skiprows=1, header=None, usecols=[0, 2, 4, 5, 6, 7, 8, 10, 11, 13, 14, 16, 18, 19, 21, 22], 
                          names=['DIV', 'CUST_NBR', 'CUST_PO', 'GREEN_BAR', 'ORDER_DATE', 'START_SHIP', 'CXL_SHIP', 'SEASON', 'STYLE', 'COLOR', 'SP', 'ORDERED_UNITS', 'SHIPPED_UNITS', 'SHIP_ON_DATE', 'PICK_UNITS', 'UNCONFIRMED'],
                          converters={'DIV':np.int16, 'CUST_NBR': str, 'CUST_PO': str, 'GREEN_BAR': str, 'ORDER_DATE': str, 'START_SHIP': str, 'CXL_SHIP': str, 'SEASON': str, 'STYLE': str, 'COLOR': str, 'SP': np.float64, 'ORDERED_UNITS': np.int64, 'SHIPPED_UNITS': np.int64, 'SHIP_ON_DATE': str, 'PICK_UNITS': np.int64, 'UNCONFIRMED': str})
    
    pdOrder['CUST_NBR'] = pdOrder['CUST_NBR'].apply(lambda x: x.zfill(5))
    pdOrder['CUST_PO'] = pdOrder['CUST_PO'].apply(lambda x: x.zfill(6) if len(x) < 6 else x)
   
#    pdOrder['ORDER_DATE'] = pd.to_datetime(pdOrder['ORDER_DATE'], errors='coerce')
#    pdOrder['START_SHIP'] = pd.to_datetime(pdOrder['START_SHIP'], errors='coerce')
#    pdOrder['CXL_SHIP'] = pd.to_datetime(pdOrder['CXL_SHIP'], errors='coerce')
#    pdOrder['SHIP_ON_DATE'] = pd.to_datetime(pdOrder['SHIP_ON_DATE'], errors='coerce')
    pdOrder['ORDER_DATE'] = pdOrder['ORDER_DATE'].apply(lambda x: dt.strptime(x[:10], '%Y-%m-%d').date())
    pdOrder['START_SHIP'] = pdOrder['START_SHIP'].apply(lambda x: dt.strptime(x[:10], '%Y-%m-%d').date())
    pdOrder['CXL_SHIP'] = pdOrder['CXL_SHIP'].apply(lambda x: dt.strptime(x[:10], '%Y-%m-%d').date())
    pdOrder['SHIP_ON_DATE'] = pdOrder['SHIP_ON_DATE'].apply(lambda x: dt.strptime(x[:10], '%Y-%m-%d').date() if pd.isnull(x) == False else pd.NaT)
    
    pdOrder['SEASON'] = pdOrder['SEASON'].apply(lambda x: x[:2].upper() + '-' + x[-2:])
    pdOrder['UNCONFIRMED'] = pdOrder['UNCONFIRMED'].apply(lambda x: 1 if x == '*' else 0)
    pdOrder['STYLE'] = pdOrder['STYLE'].apply(lambda x: x.zfill(6) if len(x) < 6 else x)
    pdOrder['COLOR'] = pdOrder['COLOR'].apply(lambda x: x.zfill(3))   
    
    ship_date_by_po = pd.pivot_table(pdOrder, values=['SHIPPED_UNITS'], index=['GREEN_BAR', 'SHIP_ON_DATE'], aggfunc='sum').reset_index()
    pdOrder = pd.merge(pdOrder, ship_date_by_po[['GREEN_BAR', 'SHIP_ON_DATE']], how='left', on=['GREEN_BAR'])
    pdOrder.drop(columns=['SHIP_ON_DATE_x'], inplace=True)
    pdOrder.rename(index=str, columns={'SHIP_ON_DATE_y': 'SHIP_ON_DATE'}, inplace=True)
    return pdOrder

def multiSeasonOrder ():
#    fileList = [each for each in os.listdir() if each[:17] == 'SOURCE_CUST_ORDER']
#    if len(fileList) >0:
#        for n in range(len(fileList)):
#            if n == 0:
#                multiSeason = oneSeasonOrder(fileList[n])
#            else:
#                addone = oneSeasonOrder(fileList[n])
#                multiSeason = multiSeason.append(addone)
#    wb = load_workbook(sourcefile, read_only=True)
#    orderTabs = [each for each in wb.sheetnames if 'ALL' in each]
    if len(orderTabs) >0:
        for n in range(len(orderTabs)):
            if n == 0:
                multiSeason = oneSeasonOrder(sourcefile, orderTabs[n])
            else:
                addone = oneSeasonOrder(sourcefile, orderTabs[n])
                multiSeason = multiSeason.append(addone)
#        pdOO = pd.read_csv('SOURCE_OO.csv', skiprows=1, header=None, usecols=[4, 10, 12, 16, 17, 19, 20, 21, 22, 23, 24, 25, 26], 
#                   names=['GREEN_BAR', 'STYLE', 'COLOR', 'SHIPPED', 'BAL', 'S1', 'S2', 'S3', 'S4', 'S5', 'S6', 'S7', 'S8'], converters={4: str, 10: str, 12: str})
        try:
            pdOO = pd.read_excel(sourcefile, sheet_name='OO', skiprows=1, header=None, usecols=[4, 10, 12, 16, 17, 19, 20, 21, 22, 23, 24, 25, 26], 
                                 names=['GREEN_BAR', 'STYLE', 'COLOR', 'SHIPPED', 'BAL', 'S1', 'S2', 'S3', 'S4', 'S5', 'S6', 'S7', 'S8'], 
                                 converters={'GREEN_BAR': str, 'STYLE': str, 'COLOR': str})
        except:
            sys.exit('There is no "OO" file to read')
        pdOO['STYLE'] = pdOO['STYLE'].apply(lambda x: x.zfill(6) if len(x) < 6 else x)
        pdOO['COLOR'] = pdOO['COLOR'].apply(lambda x: x.zfill(3))
        newMulti = pd.merge(multiSeason, pdOO[['GREEN_BAR', 'STYLE', 'COLOR', 'BAL']], how='left', on=['GREEN_BAR', 'STYLE', 'COLOR'])
        newMulti['BAL'].fillna(0, inplace=True)
        """
        Orders are categorized below:
            DEACTIVE are orders that have been cancelled but are still showing in the data pulled (*not sure how it's different from orders no longer showing up)
                update all the balances to 0 (size break)
            COMPLETE are orders that we have shipped in full
                update all the balances to 0 (size break)
            PARTIAL are orders that we have shipped but have balances left
                use OO to update the balances but not the orders (size break)
            UNCONFIRMED are orders that are not confirmed (i.e. in HFC)
                use OO to update both balances and orders (size break)
            ACTIVE are customer PO orders that are in the pipeline waiting to ship
                use OO to update both balances and orders (size break)
        """
        newMulti['COMMENT'] = newMulti.apply(lambda row: 'DEACTIVE' if ((row.SHIPPED_UNITS == 0) & (pd.isnull(row.SHIP_ON_DATE) == False) & (row.BAL == 0))
                else 'COMPLETE' if ((row.SHIPPED_UNITS >= row.ORDERED_UNITS) & (pd.isnull(row.SHIP_ON_DATE) == False))
                else 'PARTIAL' if ((row.BAL > 0) & (pd.isnull(row.SHIP_ON_DATE) == False))
                else 'UNCONFIRMED' if row.UNCONFIRMED == 1
                else 'ACTIVE', axis=1)
        
        pdOrder = pdOO[(pdOO['SHIPPED'] == 0) & (pdOO['BAL'] > 0)]
        pdOrder = pdOrder[['GREEN_BAR', 'STYLE', 'COLOR', 'S1', 'S2', 'S3', 'S4', 'S5', 'S6', 'S7', 'S8']]
        
        pdPartial = pdOO[(pdOO['SHIPPED'] > 0) & (pdOO['BAL'] > 0)]
        pdPartial = pdPartial[['GREEN_BAR', 'STYLE', 'COLOR', 'S1', 'S2', 'S3', 'S4', 'S5', 'S6', 'S7', 'S8']]

        newMulti.to_sql('#temp_cust_order', con=conn, if_exists='replace', index=False)
        pdOrder.to_sql('#temp_full_order', con=conn, if_exists='replace', index=False)
        pdPartial.to_sql('#temp_partial_order', con=conn, if_exists='replace', index=False)
        trans = conn.begin()
        try:
            conn.execute("""MERGE DBO.CUST_ORDER AS T 
                         USING dbo.#temp_cust_order AS S 
                         ON (T.GREEN_BAR = S.GREEN_BAR and T.STYLE = S.STYLE and T.COLOR = S.COLOR)
                         WHEN MATCHED THEN UPDATE
                         SET T.DIV=S.DIV, T.CUST_NBR=S.CUST_NBR, T.ORDER_DATE=S.ORDER_DATE, T.START_SHIP=S.START_SHIP, T.CXL_SHIP=S.CXL_SHIP, T.CUST_PO=S.CUST_PO, 
                         T.SEASON=S.SEASON, T.SP=S.SP, T.ORDERED_UNITS=S.ORDERED_UNITS, T.SHIPPED_UNITS=S.SHIPPED_UNITS, T.BAL_UNITS=S.BAL, T.PICK_UNITS=S.PICK_UNITS, 
                         T.SHIP_ON_DATE = S.SHIP_ON_DATE, T.UNCONFIRMED = S.UNCONFIRMED, T.COMMENT=S.COMMENT, T.CXL=0 
                         WHEN NOT MATCHED BY TARGET THEN 
                         INSERT (GREEN_BAR, STYLE, COLOR, DIV, CUST_NBR, ORDER_DATE, START_SHIP, CXL_SHIP, CUST_PO, SEASON, SP, ORDERED_UNITS, SHIPPED_UNITS, BAL_UNITS, PICK_UNITS, SHIP_ON_DATE, UNCONFIRMED, COMMENT) VALUES
                         (S.GREEN_BAR, S.STYLE, S.COLOR, S.DIV, S.CUST_NBR, S.ORDER_DATE, S.START_SHIP, S.CXL_SHIP, S.CUST_PO, S.SEASON, S.SP, S.ORDERED_UNITS, S.SHIPPED_UNITS, S.BAL, S.PICK_UNITS, S.SHIP_ON_DATE, S.UNCONFIRMED, S.COMMENT)
                         WHEN NOT MATCHED BY SOURCE AND (T.SEASON IN (select distinct SEASON from #temp_cust_order)) THEN
                         UPDATE SET T.CXL=1, T.BAL_UNITS = 0, T.BAL_S1 = 0, T.BAL_S2 = 0, T.BAL_S3 = 0, T.BAL_S4 = 0, T.BAL_S5 = 0, T.BAL_S6 = 0, T.BAL_S7 = 0, T.BAL_S8 =0;""")
            conn.execute("""UPDATE T
                         SET T.ORDERED_S1 = S.S1, T.ORDERED_S2 = S.S2, T.ORDERED_S3 = S.S3, T.ORDERED_S4 = S.S4, T.ORDERED_S5 = S.S5, T.ORDERED_S6 = S.S6, T.ORDERED_S7 = S.S7, T.ORDERED_S8 = S.S8, 
                         T.BAL_S1 = S.S1, T.BAL_S2 = S.S2, T.BAL_S3 = S.S3, T.BAL_S4 = S.S4, T.BAL_S5 = S.S5, T.BAL_S6 = S.S6, T.BAL_S7 = S.S7, T.BAL_S8 = S.S8
                         FROM DBO.CUST_ORDER AS T INNER JOIN #temp_full_order AS S ON (S.GREEN_BAR = T.GREEN_BAR and S.STYLE = T.STYLE and S.COLOR = T.COLOR);""") 
            conn.execute("""UPDATE T
                         SET T.BAL_S1 = S.S1, T.BAL_S2 = S.S2, T.BAL_S3 = S.S3, T.BAL_S4 = S.S4, T.BAL_S5 = S.S5, T.BAL_S6 = S.S6, T.BAL_S7 = S.S7, T.BAL_S8 = S.S8
                         FROM DBO.CUST_ORDER AS T INNER JOIN #temp_partial_order AS S ON (S.GREEN_BAR = T.GREEN_BAR and S.STYLE = T.STYLE and S.COLOR = T.COLOR);""")
            conn.execute("""UPDATE T
                         SET T.BAL_S1 = 0, T.BAL_S2 = 0, T.BAL_S3 = 0, T.BAL_S4 = 0, T.BAL_S5 = 0, T.BAL_S6 = 0, T.BAL_S7 = 0, T.BAL_S8 = 0
                         FROM DBO.CUST_ORDER AS T WHERE T.COMMENT IN ('COMPLETE', 'DEACTIVE');""")
            trans.commit()
            print ('CUST_ORDER TABLE UPDATE COMPLETED SUCCESSFULLY FROM SOURCE ' + sourcefile.split("\\")[-1])
        except Exception as e:
            logger = logging.Logger('Catch_All')
            logger.error(str(e))
            trans.rollback()
            conn.close()
            engine.dispose()
            sys.exit('PROGRAM STOPPED #1')
        
        ### BELOW BLOCK OF CODE IS ADDED TO AUTO DETERMINE IF A VALID ORDER (i.e. any order that is not purged nor de-activated) is an UPFRONT order or OFFPRICE order
        ####################### CODE BEGINS #########################################
        sql_identify_offprice = '''
                                SELECT C.GREEN_BAR, C.STYLE, C.COLOR, C.CUST_NBR, C.SEASON AS ORDER_SSN, S.SEASON AS STYLE_SSN, T.SEASON AS HFC_SSN FROM DBO.CUST_ORDER C
                                LEFT JOIN (SELECT STYLE, SEASON FROM DBO.STYLE_MASTER) AS S ON C.STYLE = S.STYLE
                                LEFT JOIN (SELECT DISTINCT D.STYLE, H.CUST_NBR, H.SEASON FROM DBO.HFC_DETAIL D JOIN DBO.HFC_HEADER H ON D.HFC_NBR = H.HFC_NBR WHERE D.CXL = 0) AS T ON C.STYLE = T.STYLE AND C.CUST_NBR = T.CUST_NBR
                                WHERE (C.CXL = 0 AND C.COMMENT <> 'DEACTIVE');
                                '''
        pd_identify = pd.read_sql(sql_identify_offprice, con=conn)
        """
        OFFPRICE order selection criteria #1: if HFC_SSN is null, it is an OFFPRICE order; UNLESS
        EXCEPTION: When the customer is Haggar (42070 & 42071 interchangeable) or Shopko (78188 & 78190 interchangeable)
                    This works because we know Haggar and Shopko do not buy close-out orders; If in future this condition changes, we need to reconsider below code
        """
        pd_identify_3 = pd_identify[(pd_identify['HFC_SSN'].isnull()) & ((pd_identify['CUST_NBR'] != '42071') & (pd_identify['CUST_NBR'] != '42070') & (pd_identify['CUST_NBR'] != '78190'))]
        """
        OFFPRICE order selection criteria #2: if HFC_SSN is not null, and ORDER_SSN and HFC_SSN are different, then it is an OFFPRICE order; UNLESS
        EXCEPTION: We have known situations where we bring Haggar orders under a different season code than its HFC (because they mix different season HFC on one PO)
                    We need to manually add those Haggar PO numbers (CUST_PO) in the SQL line below to change them from OFFPRICE to UPFRONT
                   ***also, originally thought about giving this exception to Amazon too (we ship older season styles to them) but changed my mind. OK to let those be labelled OFFPRICE
        """
        pd_identify_2 = pd_identify[(pd_identify['ORDER_SSN'] != pd_identify['HFC_SSN']) & (~pd_identify['HFC_SSN'].isnull())]
        pd_identify_2 = pd_identify_2.append(pd_identify_3)
        pd_identify_2['COMMENT'] = 'OFFPRICE'
        pd_identify = pd.merge(pd_identify, pd_identify_2[['GREEN_BAR', 'STYLE', 'COLOR', 'COMMENT']], how='left', on=['GREEN_BAR', 'STYLE', 'COLOR'])
        pd_identify['COMMENT'].fillna('UPFRONT', inplace=True)
        
        pd_identify[['GREEN_BAR', 'STYLE', 'COLOR', 'COMMENT']].to_sql('#temp_order_comment', con=conn, if_exists='replace', index=False)
        trans = conn.begin()
        try:
            conn.execute("""UPDATE T SET T.COMMENT_2 = S.COMMENT FROM DBO.CUST_ORDER AS T INNER JOIN #temp_order_comment AS S 
                         ON (T.GREEN_BAR = S.GREEN_BAR and T.STYLE = S.STYLE and T.COLOR = S.COLOR);""")
            conn.execute("""UPDATE dbo.CUST_ORDER SET COMMENT_2 = 'UPFRONT' WHERE CUST_PO IN ('114350', '1037316', '115040');""")  ### this line was added to deal with the previously mentioned exception of Div10 Haggar PO's brought in under different season code
            trans.commit()
            print ('OFFPRICE and UPFRONT identifications were loaded successfully to CUST_ORDER table')
        except Exception as e:
            logger = logging.Logger('Catch_All')
            logger.error(str(e))
            trans.rollback()
            conn.close()
            engine.dispose()
            sys.exit('PROGRAM STOPPED #2')
        ###################### CODE ENDS ############################################       
    else:
        sys.exit('There is no customer order file to read')

def oneSeason (aFile, aTab):
    assoc = pd.read_excel(aFile, sheet_name=aTab, skiprows=1, header=None, names=['GREEN_BAR', 'STYLE', 'COLOR', 'HFC'], usecols=[1, 3, 4, 5], converters={'GREEN_BAR': str, 'STYLE': str, 'COLOR':str, 'HFC':str})
    assoc['STYLE'] = assoc['STYLE'].apply(lambda x: x.zfill(6) if len(x) < 6 else x)
    assoc['COLOR'] = assoc['COLOR'].apply(lambda x: x.zfill(3))
    assoc['HFC'] = assoc['HFC'].apply(lambda x: x.split('-')[0].zfill(6) if pd.isnull(x) == False else '')
    assoc['HFC'] = assoc['HFC'].apply(lambda x: x if len(x) == 6 else '')
    assoc['SEASON'] = aTab[:5].upper()    
    return assoc

def allSeasonAssoc ():
#    fileList = [each for each in os.listdir() if each[:16] == 'SOURCE_HFC_ASSOC']
    if len(assocTabs) >0:
        for n in range(len(assocTabs)):
            if n == 0:
                allAssoc = oneSeason(sourcefile, assocTabs[n])
            else:
                addone = oneSeason(sourcefile, assocTabs[n])
                allAssoc = allAssoc.append(addone)
        allAssoc.to_sql('#temp_hfc_assoc', con=conn, if_exists='replace', index=False)
        trans = conn.begin()
        try:
            conn.execute("""MERGE DBO.HFC_ASSOC AS T 
                         USING dbo.#temp_hfc_assoc AS S 
                         ON (T.GREEN_BAR = S.GREEN_BAR and T.STYLE=S.STYLE and T.COLOR=S.COLOR) 
                         WHEN MATCHED THEN UPDATE
                         SET T.HFC=S.HFC, T.SEASON=S.SEASON, T.CXL=0
                         WHEN NOT MATCHED BY TARGET THEN 
                         INSERT (GREEN_BAR, STYLE, COLOR, HFC, SEASON) VALUES
                         (S.GREEN_BAR, S.STYLE, S.COLOR, S.HFC, S.SEASON)
                         WHEN NOT MATCHED BY SOURCE THEN
                         UPDATE SET T.CXL=1;""")
            # Below section of codes are updated throughout a season to correct any HFC attachment issues in PROCOMM # 
            conn.execute ("""update dbo.HFC_ASSOC set HFC = '197218' where GREEN_BAR = '878294' AND STYLE = '138196' AND COLOR = '503';""")
            conn.execute ("""update dbo.HFC_ASSOC set CXL = 0 where GREEN_BAR = '876766';""")
            # Above section of codes are updated throughout a season to correct any HFC attachment issues in PROCOMM #
            trans.commit()         
            print ('HFC_ASSOC TABLE UPDATE COMPLETED SUCCESSFULLY FROM SOURCE ' + sourcefile.split("\\")[-1])
        except Exception as e:
            logger = logging.Logger('Catch_All')
            logger.error(str(e))
            trans.rollback()
            conn.close()
            engine.dispose()
            sys.exit('PROGRAM STOPPED #3')
    else:
        sys.exit('There is no association file to read')

def uploadInv ():
    inv = pd.read_excel(sourcefile, invTab, skiprows=1, header=None, usecols=[7,8,9,14,19,20,21,22,23,24,25,26], names=['STYLE', 'COLOR_DESC', 'COLOR_CODE', 'OH', 'S1', 'S2', 'S3', 'S4', 'S5', 'S6', 'S7', 'S8'],
                  converters = {'STYLE': str, 'COLOR_CODE': str, 'OH': np.int32, 'S1': np.int32, 'S2': np.int32, 'S3':np.int32, 'S4':np.int32, 'S5':np.int32, 'S6':np.int32, 'S7':np.int32, 'S8':np.int32})
    inv['STYLE'] = inv['STYLE'].apply(lambda x: x.zfill(6))
    inv['COLOR_CODE'] = inv['COLOR_CODE'].apply(lambda x: x.zfill(3))
    
    bal_by_style_by_color = '''SELECT STYLE, COLOR as COLOR_CODE, SUM(BAL_UNITS) AS TTL_BAL, SUM(BAL_S1) AS O1, SUM(BAL_S2) AS O2, SUM(BAL_S3) AS O3, SUM(BAL_S4) AS O4,
                            SUM(BAL_S5) AS O5, SUM(BAL_S6) AS O6, SUM(BAL_S7) AS O7, SUM(BAL_S8) AS O8 FROM DBO.CUST_ORDER WHERE CXL = 0 GROUP BY STYLE, COLOR'''
    bal = pd.read_sql(bal_by_style_by_color, con=conn)
    
    inv = pd.merge(inv, bal, how='left', on=['STYLE', 'COLOR_CODE']).fillna(0)
    # this should work after new season (spring 2020) is uploaded into database
#    inv['OH_S1'] = inv['S1'] + inv['O1']
#    inv['OH_S2'] = inv['S2'] + inv['O2']
#    inv['OH_S3'] = inv['S3'] + inv['O3']
#    inv['OH_S4'] = inv['S4'] + inv['O4']
#    inv['OH_S5'] = inv['S5'] + inv['O5']
#    inv['OH_S6'] = inv['S6'] + inv['O6']
#    inv['OH_S7'] = inv['S7'] + inv['O7']
#    inv['OH_S8'] = inv['S8'] + inv['O8']
    inv['OH_S1'] = inv.apply(lambda row: row.S1 + row.O1 if (row.S1 + row.O1) > 0 else 0, axis=1)
    inv['OH_S2'] = inv.apply(lambda row: row.S2 + row.O2 if (row.S2 + row.O2) > 0 else 0, axis=1)
    inv['OH_S3'] = inv.apply(lambda row: row.S3 + row.O3 if (row.S3 + row.O3) > 0 else 0, axis=1)
    inv['OH_S4'] = inv.apply(lambda row: row.S4 + row.O4 if (row.S4 + row.O4) > 0 else 0, axis=1)
    inv['OH_S5'] = inv.apply(lambda row: row.S5 + row.O5 if (row.S5 + row.O5) > 0 else 0, axis=1)
    inv['OH_S6'] = inv.apply(lambda row: row.S6 + row.O6 if (row.S6 + row.O6) > 0 else 0, axis=1)
    inv['OH_S7'] = inv.apply(lambda row: row.S7 + row.O7 if (row.S7 + row.O7) > 0 else 0, axis=1)
    inv['OH_S8'] = inv.apply(lambda row: row.S8 + row.O8 if (row.S8 + row.O8) > 0 else 0, axis=1)
    inv['OH_CAL'] = inv['OH_S1'] + inv['OH_S2'] +inv['OH_S3'] +inv['OH_S4'] +inv['OH_S5'] +inv['OH_S6'] +inv['OH_S7'] +inv['OH_S8']
    inv['Diff'] = inv['OH_CAL'] - inv['OH']
    exception = inv[inv['Diff'] != 0]
    if exception.shape[0] > 0:
        exception.to_excel('ExceptionReport_ORDER_AND_INV.xlsx', index=False)
        choice = str(input("Please review offloaded ExceptionReport_ORDER_AND_INV.xlsx. IS IT OKAY TO PROCEED? ENTER 'Y' OR 'N'"))
        if choice.upper() == 'Y':
            inv_upload = inv[['STYLE', 'COLOR_CODE', 'COLOR_DESC', 'OH', 'OH_S1', 'OH_S2', 'OH_S3', 'OH_S4', 'OH_S5', 'OH_S6', 'OH_S7', 'OH_S8']]
        else:
            conn.close()
            engine.dispose()
            sys.exit('PROGRAM STOPPED #4')
    else:
        print('Inventory data review passed! Proceed to uploading...')
        inv_upload = inv[['STYLE', 'COLOR_CODE', 'COLOR_DESC', 'OH', 'OH_S1', 'OH_S2', 'OH_S3', 'OH_S4', 'OH_S5', 'OH_S6', 'OH_S7', 'OH_S8']]
    
    inv_upload.to_sql('#temp_inv', con=conn, if_exists='replace', index=False)
    trans = conn.begin()
    try:
        conn.execute("""MERGE DBO.INV AS T 
                     USING dbo.#temp_inv AS S 
                     ON (T.STYLE=S.STYLE and T.COLOR_CODE=S.COLOR_CODE) 
                     WHEN MATCHED THEN UPDATE
                     SET T.OH = S.OH, T.OH_S1=S.OH_S1, T.OH_S2=S.OH_S2, T.OH_S3=S.OH_S3, T.OH_S4=S.OH_S4,T.OH_S5=S.OH_S5,T.OH_S6=S.OH_S6,T.OH_S7=S.OH_S7,T.OH_S8=S.OH_S8, T.CXL = 0
                     WHEN NOT MATCHED BY TARGET THEN 
                     INSERT (STYLE, COLOR_CODE, COLOR_DESC, OH, OH_S1, OH_S2,OH_S3,OH_S4,OH_S5,OH_S6,OH_S7,OH_S8) VALUES
                     (S.STYLE, S.COLOR_CODE, S.COLOR_DESC, S.OH, S.OH_S1, S.OH_S2, S.OH_S3, S.OH_S4, S.OH_S5, S.OH_S6, S.OH_S7,S.OH_S8)
                     WHEN NOT MATCHED BY SOURCE THEN
                     UPDATE SET T.OH = 0, T.OH_S1=0, T.OH_S2=0, T.OH_S3=0, T.OH_S4=0, T.OH_S5=0, T.OH_S6=0, T.OH_S7=0, T.OH_S8=0, T.CXL=1;""")
        trans.commit()         
        print ('INV TABLE UPDATE COMPLETED SUCCESSFULLY FROM SOURCE ' + sourcefile.split("\\")[-1])
    except Exception as e:
        logger = logging.Logger('Catch_All')
        logger.error(str(e))
        trans.rollback()
        conn.close()
        engine.dispose()
        sys.exit('PROGRAM STOPPED #5')          
    

multiSeasonOrder()
allSeasonAssoc()  # This module has lines added in throughout the season to manually fix HFC attachment issues in PROCOMM; update those lines as necessary
uploadInv()

conn.close()
engine.dispose()
